#!/usr/bin/env python3
"""
xlsx_tool.py - Read and edit Excel workbooks.

Read-only operations use openpyxl by default because they do not save or
rewrite the workbook. Cell edits default to a direct OOXML writer that modifies
only the required package parts and preserves unsupported workbook features by
leaving unrelated XML untouched. Excel COM remains available for live
workbooks, formula validation, sheet copying, and formula-aware structural
operations.

Usage:
    python tools/xlsx/xlsx_tool.py read <file> [--sheet NAME] [--range A1:Z100]
    python tools/xlsx/xlsx_tool.py list-sheets <file>
    python tools/xlsx/xlsx_tool.py search <file> --query TEXT [--sheet NAME]
    python tools/xlsx/xlsx_tool.py edit <file> --sheet NAME --cell A1 --value TEXT
    python tools/xlsx/xlsx_tool.py edit <file> --sheet NAME --edits '[{"cell":"A1","value":"x"}]'
    python tools/xlsx/xlsx_tool.py edit <file> --edits-file edits.json
    python tools/xlsx/xlsx_tool.py add-sheet <file> "Support"
    python tools/xlsx/xlsx_tool.py move-sheet <file> "Support" --before "Divider"
    python tools/xlsx/xlsx_tool.py insert <file> --sheet "Sheet1" --axis columns --range D:D
    python tools/xlsx/xlsx_tool.py validate <file> --full-calc --save
    python tools/xlsx/xlsx_tool.py copy-sheets <file> --source <source.xlsx> --sheets "Sheet1" "Old=New"

For COM operations, if the target workbook is already open in Excel, the tool
attaches to that live workbook and leaves it open. Otherwise it uses a temporary
hidden Excel instance and closes it after the command.
"""

# === office-tools bootstrap ============================================
# Ensure third-party deps are installed in the plugin's persistent data dir
# before any third-party import runs. Idempotent and cheap after the first
# install. Adds CLAUDE_PLUGIN_DATA/site-packages to sys.path either way.
import os
import sys

_PLUGIN_ROOT = os.environ.get("CLAUDE_PLUGIN_ROOT") or os.environ.get("PLUGIN_ROOT")
if _PLUGIN_ROOT and _PLUGIN_ROOT not in sys.path:
    sys.path.insert(0, _PLUGIN_ROOT)
else:
    # Dev fallback: tool lives at <plugin_root>/tools/<name>/<file>.py
    sys.path.insert(
        0,
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    )

from bootstrap.ensure_python_deps import run as _ensure_python_deps  # noqa: E402

_ensure_python_deps()
# === end bootstrap =====================================================

import argparse
import io
import json
import posixpath
import re
import shutil
import tempfile
import zipfile
from datetime import datetime, date, time
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
import xlwings as xw

xw.LICENSE_KEY = "noncommercial"


def format_cell_value(value):
    """Format a cell value for display."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return f"{value:.2f}"
    return str(value)


def parse_excel_datetime(value):
    """Parse common ISO-like date inputs for Excel COM writes."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time())
    if not isinstance(value, str):
        return value

    text = value.strip()
    if not text:
        return value

    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return value


def coerce_edit_value(value, value_type=None):
    """Coerce JSON/CLI edit values before assigning through Excel COM."""
    if value_type:
        requested = str(value_type).lower()
        if requested in {"text", "str", "string"}:
            return "" if value is None else str(value)
        if requested in {"date", "datetime"}:
            return parse_excel_datetime(value)
        if requested in {"int", "integer"}:
            return int(value)
        if requested in {"float", "number", "numeric"}:
            return float(value)
        print(f"Error: Unsupported value type: {value_type}")
        sys.exit(1)

    # Auto-detect formulas, ISO-like dates, and numeric types from string values.
    if isinstance(value, str) and not value.startswith("="):
        parsed_date = parse_excel_datetime(value)
        if parsed_date is not value:
            return parsed_date
        try:
            return int(value)
        except ValueError:
            try:
                return float(value)
            except ValueError:
                pass
    return value


def excel_address(range_api):
    """Return an A1 address without $ markers for an Excel COM range."""
    try:
        return range_api.Address(False, False)
    except Exception:
        try:
            return range_api.Address
        except Exception:
            return "<unknown>"


def iter_locked_cells(range_api, limit=20):
    """Return up to limit locked cell addresses from an Excel COM range."""
    locked = []
    try:
        cells = range_api.Cells
        count = int(getattr(cells, "CountLarge", cells.Count))
    except Exception:
        count = 0

    if count <= 0:
        return locked

    for index in range(1, count + 1):
        try:
            cell = cells.Item(index)
            if bool(cell.Locked):
                locked.append(excel_address(cell))
                if len(locked) >= limit:
                    break
        except Exception:
            continue
    return locked


def assert_editable_range(ws, target, cell_ref):
    """Fail before editing locked cells on a protected sheet."""
    try:
        protected = bool(target.api.Worksheet.ProtectContents)
    except Exception:
        protected = False
    if not protected:
        return

    check_api = target.api
    try:
        if bool(target.api.MergeCells):
            check_api = target.api.MergeArea
    except Exception:
        pass

    try:
        locked_state = check_api.Locked
    except Exception:
        locked_state = None

    if locked_state is False:
        return

    if locked_state is True:
        address = excel_address(check_api)
        raise RuntimeError(
            f"Protected locked cell edit blocked: {ws.name}!{address}. "
            "The sheet is protected and the target range is locked; write to "
            "an unlocked input/comment cell or a separate support sheet."
        )

    locked_cells = iter_locked_cells(check_api)
    if locked_cells:
        more = ""
        try:
            count = int(getattr(check_api.Cells, "CountLarge", check_api.Cells.Count))
            if count > len(locked_cells):
                more = f" (showing first {len(locked_cells)} locked cells)"
        except Exception:
            pass
        raise RuntimeError(
            f"Protected locked cell edit blocked: {ws.name}!{cell_ref} includes "
            f"locked cell(s) {', '.join(locked_cells)}{more}. The sheet is "
            "protected; write only to unlocked input/comment cells or a separate "
            "support sheet."
        )


def clear_contents_com(target):
    """Clear a COM range, falling back to merge-area clears when needed."""
    try:
        target.api.ClearContents()
        return
    except Exception as first_error:
        try:
            if bool(target.api.MergeCells):
                target.api.MergeArea.ClearContents()
                return
        except Exception:
            pass

        try:
            cells = target.api.Cells
            count = int(getattr(cells, "CountLarge", cells.Count))
        except Exception:
            raise first_error

        if count > 200000:
            raise RuntimeError(
                f"Range clear failed and fallback would scan {count} cells. "
                "Narrow the clear range or use the direct engine."
            ) from first_error

        cleared_merges = set()
        for index in range(1, count + 1):
            cell = cells.Item(index)
            try:
                if bool(cell.MergeCells):
                    area = cell.MergeArea
                    address = area.Address(False, False)
                    if address not in cleared_merges:
                        area.ClearContents()
                        cleared_merges.add(address)
                else:
                    cell.ClearContents()
            except Exception:
                raise first_error


def range_values_as_rows(range_obj):
    """Return an xlwings range value as a consistent list of row lists."""
    raw = range_obj.value
    if raw is None:
        return []

    row_count = range_obj.rows.count
    col_count = range_obj.columns.count

    if row_count == 1 and col_count == 1:
        return [[raw]]
    if row_count == 1:
        if isinstance(raw, list):
            return [raw]
        return [[raw]]
    if col_count == 1:
        if isinstance(raw, list):
            return [[value] for value in raw]
        return [[raw]]
    if not isinstance(raw, list):
        return [[raw]]
    if raw and not isinstance(raw[0], list):
        return [raw]
    return raw


def load_readonly_workbook(filepath, data_only=True):
    """Open a workbook for read-only inspection without saving it."""
    try:
        return load_workbook(
            filepath,
            read_only=True,
            data_only=data_only,
            keep_links=False,
        )
    except Exception as e:
        print(f"Error: Could not read workbook with openpyxl: {e}")
        sys.exit(1)


def openpyxl_dimensions(ws):
    """Return a sheet dimension string, handling empty sheets consistently."""
    try:
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            return "A1"
    except Exception:
        pass
    return ws.calculate_dimension()


def openpyxl_values_as_rows(ws, range_ref=None):
    """Return worksheet values as row lists for a range or used dimensions."""
    if range_ref:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    else:
        if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
            return []
        min_row = 1
        min_col = 1
        max_row = ws.max_row
        max_col = ws.max_column

    rows = []
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
        values_only=True,
    ):
        rows.append(list(row))
    return rows


def print_markdown_table(rows):
    """Print worksheet rows as a compact markdown table."""
    data = [[format_cell_value(v) for v in row] for row in rows]
    if not data:
        print("(no data)")
        return

    max_cols = max(len(r) for r in data)
    for r in data:
        while len(r) < max_cols:
            r.append("")

    col_widths = [0] * max_cols
    for row in data:
        for i, val in enumerate(row):
            col_widths[i] = max(col_widths[i], len(val))
    col_widths = [min(max(w, 3), 50) for w in col_widths]

    header = data[0]
    print("| " + " | ".join(
        val.ljust(col_widths[i])[:col_widths[i]] for i, val in enumerate(header)
    ) + " |")
    print("| " + " | ".join("-" * col_widths[i] for i in range(max_cols)) + " |")

    for row in data[1:]:
        print("| " + " | ".join(
            val.ljust(col_widths[i])[:col_widths[i]] for i, val in enumerate(row)
        ) + " |")


def is_blank_cell_value(value):
    return value is None or value == ""


def compact_display_rows(rows, show_empty_rows=False, show_empty_cols=False):
    """Drop fully empty rows/columns for readable CLI output."""
    if not rows:
        return []

    filtered = rows if show_empty_rows else [
        row for row in rows if any(not is_blank_cell_value(value) for value in row)
    ]
    if not filtered or show_empty_cols:
        return filtered

    max_cols = max(len(row) for row in filtered)
    keep_cols = []
    for col_idx in range(max_cols):
        if any(
            col_idx < len(row) and not is_blank_cell_value(row[col_idx])
            for row in filtered
        ):
            keep_cols.append(col_idx)

    if not keep_cols:
        return []

    return [[row[idx] if idx < len(row) else None for idx in keep_cols] for row in filtered]


def normalise_path(path):
    """Return a comparable absolute path for a workbook."""
    return os.path.normcase(os.path.abspath(os.path.realpath(path)))


def find_open_book(filepath):
    """Return an already-open workbook matching filepath, if one exists."""
    target = normalise_path(filepath)

    try:
        apps = list(xw.apps)
    except Exception:
        return None, None

    for app in apps:
        try:
            books = list(app.books)
        except Exception:
            continue

        for book in books:
            try:
                if normalise_path(book.fullname) == target:
                    return app, book
            except Exception:
                continue

    return None, None


def open_book(filepath, readonly=False, prefer_open=True, keep_open=False):
    """Open a workbook, preferring an already-open Excel workbook."""
    if prefer_open:
        app, wb = find_open_book(filepath)
        if wb is not None:
            return app, wb, False

    app = xw.App(visible=keep_open, add_book=False)
    if not keep_open:
        app.display_alerts = False
        app.screen_updating = False
        try:
            app.calculation = "manual"
        except Exception:
            pass
        app.api.AskToUpdateLinks = False

    open_kwargs = {"update_links": False}
    if readonly:
        open_kwargs["read_only"] = True

    try:
        wb = app.books.open(filepath, **open_kwargs)
    except Exception:
        try:
            app.quit()
        except Exception:
            pass
        raise
    return app, wb, not keep_open


def close_book(app, wb, save=False, close_on_exit=True):
    """Save and close only workbooks opened for this command."""
    try:
        if save:
            wb.save()
        if close_on_exit:
            wb.close()
    finally:
        if close_on_exit:
            try:
                app.quit()
            except Exception:
                pass


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
XML_NS = "http://www.w3.org/XML/1998/namespace"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
CALC_CHAIN_REL_TYPE = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"

COMMON_OOXML_NAMESPACES = {
    "": MAIN_NS,
    "r": REL_NS,
    "rel": PKG_REL_NS,
    "ct": CONTENT_TYPES_NS,
    "mc": MC_NS,
    "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "x15": "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
    "xr6": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision6",
    "xr10": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision10",
    "xm": "http://schemas.microsoft.com/office/excel/2006/main",
}

KNOWN_NAMESPACE_PREFIXES = dict(COMMON_OOXML_NAMESPACES)


def register_ooxml_namespaces():
    for prefix, uri in COMMON_OOXML_NAMESPACES.items():
        ET.register_namespace(prefix, uri)


register_ooxml_namespaces()


def qn(ns, tag):
    return f"{{{ns}}}{tag}"


def read_zip_part(zf, name):
    try:
        return zf.read(name)
    except KeyError:
        raise RuntimeError(f"Workbook package part not found: {name}")


def register_namespaces_from_xml(data):
    """Preserve workbook-specific prefixes when ElementTree reserializes XML."""
    try:
        for _event, (prefix, uri) in ET.iterparse(io.BytesIO(data), events=("start-ns",)):
            if prefix is None:
                prefix = ""
            if prefix and uri:
                KNOWN_NAMESPACE_PREFIXES[prefix] = uri
                ET.register_namespace(prefix, uri)
    except ET.ParseError:
        pass


def parse_xml_part(zf, name):
    data = read_zip_part(zf, name)
    register_namespaces_from_xml(data)
    return ET.fromstring(data)


def xml_bytes(root):
    register_ooxml_namespaces()
    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return preserve_ignorable_namespace_declarations(data, root)


def preserve_ignorable_namespace_declarations(data, root):
    """Keep xmlns declarations used only as mc:Ignorable attribute values."""
    ignorable = root.attrib.get(qn(MC_NS, "Ignorable"))
    if not ignorable:
        return data

    xml_decl_end = data.find(b"?>")
    search_from = xml_decl_end + 2 if xml_decl_end != -1 else 0
    root_start = data.find(b"<", search_from)
    root_end = data.find(b">", root_start)
    if root_start == -1 or root_end == -1:
        return data

    root_tag = data[root_start:root_end]
    additions = []
    for prefix in ignorable.split():
        marker = f"xmlns:{prefix}=".encode("utf-8")
        if marker in root_tag:
            continue
        uri = KNOWN_NAMESPACE_PREFIXES.get(prefix)
        if not uri:
            raise RuntimeError(
                f"Cannot preserve mc:Ignorable prefix '{prefix}' because its namespace URI is unknown"
            )
        additions.append(f' xmlns:{prefix}="{uri}"'.encode("utf-8"))

    if not additions:
        return data

    insert_at = root_end
    if data[root_end - 1:root_end] == b"/":
        insert_at = root_end - 1
    return data[:insert_at] + b"".join(additions) + data[insert_at:]


def rewrite_zip_parts(filepath, updates, deletes=None):
    """Rewrite selected OOXML package parts while preserving all untouched parts."""
    deletes = set(deletes or [])
    fd, tmp_path = tempfile.mkstemp(
        prefix=os.path.basename(filepath) + ".",
        suffix=".tmp",
        dir=os.path.dirname(filepath) or None,
    )
    os.close(fd)
    try:
        with zipfile.ZipFile(filepath, "r") as zin, zipfile.ZipFile(
            tmp_path, "w", compression=zipfile.ZIP_DEFLATED, allowZip64=True
        ) as zout:
            written = set()
            for info in zin.infolist():
                if info.filename in deletes:
                    continue
                if info.filename in written:
                    # Avoid preserving duplicate central-directory entries.
                    continue
                data = updates.get(info.filename)
                if data is None:
                    data = zin.read(info.filename)
                zout.writestr(info, data)
                written.add(info.filename)

            for name, data in updates.items():
                if name in deletes:
                    continue
                if name in written:
                    continue
                info = zipfile.ZipInfo(name)
                info.compress_type = zipfile.ZIP_DEFLATED
                zout.writestr(info, data)
                written.add(name)
        shutil.move(tmp_path, filepath)
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


def resolve_relationship_target(source_part, target):
    if target.startswith("/"):
        return posixpath.normpath(target.lstrip("/"))
    base_dir = posixpath.dirname(source_part)
    return posixpath.normpath(posixpath.join(base_dir, target))


def relationship_target_for_part(source_part, part_name):
    base_dir = posixpath.dirname(source_part)
    return posixpath.relpath(part_name, base_dir)


def get_workbook_package_map(zf):
    workbook_root = parse_xml_part(zf, "xl/workbook.xml")
    rels_root = parse_xml_part(zf, "xl/_rels/workbook.xml.rels")
    rid_to_target = {}
    for rel in rels_root.findall(qn(PKG_REL_NS, "Relationship")):
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rid and target:
            rid_to_target[rid] = resolve_relationship_target("xl/workbook.xml", target)

    sheets_el = workbook_root.find(qn(MAIN_NS, "sheets"))
    if sheets_el is None:
        raise RuntimeError("Workbook has no <sheets> collection")

    sheets = {}
    for sheet_el in sheets_el.findall(qn(MAIN_NS, "sheet")):
        name = sheet_el.attrib.get("name")
        rid = sheet_el.attrib.get(qn(REL_NS, "id"))
        if name and rid in rid_to_target:
            sheets[name] = {
                "element": sheet_el,
                "part": rid_to_target[rid],
                "rid": rid,
            }
    return workbook_root, rels_root, sheets_el, sheets


CELL_RE = re.compile(r"^([A-Za-z]{1,3})([1-9][0-9]*)$")


def split_cell_ref(cell_ref):
    match = CELL_RE.match(cell_ref.replace("$", ""))
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    col_letters, row_text = match.groups()
    return column_index_from_string(col_letters.upper()), int(row_text)


def cell_ref(col_idx, row_idx):
    return f"{get_column_letter(col_idx)}{row_idx}"


def cell_sort_key(ref):
    col_idx, row_idx = split_cell_ref(ref)
    return row_idx, col_idx


def iter_target_cells(cell_or_range, value):
    min_col, min_row, max_col, max_row = range_boundaries(cell_or_range)
    row_count = max_row - min_row + 1
    col_count = max_col - min_col + 1

    if isinstance(value, list):
        if value and isinstance(value[0], list):
            for r_offset, row_values in enumerate(value):
                for c_offset, item in enumerate(row_values):
                    yield min_row + r_offset, min_col + c_offset, item
            return
        for c_offset, item in enumerate(value):
            yield min_row, min_col + c_offset, item
        return

    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            yield row_idx, col_idx, value


def get_direct_child(parent, tag):
    return parent.find(qn(MAIN_NS, tag))


def ensure_direct_child(parent, tag, insert_after=None):
    child = get_direct_child(parent, tag)
    if child is not None:
        return child

    child = ET.Element(qn(MAIN_NS, tag))
    if insert_after:
        after = get_direct_child(parent, insert_after)
        if after is not None:
            index = list(parent).index(after) + 1
            parent.insert(index, child)
            return child
    parent.append(child)
    return child


def get_or_create_sheet_data(ws_root):
    sheet_data = get_direct_child(ws_root, "sheetData")
    if sheet_data is not None:
        return sheet_data

    sheet_data = ET.Element(qn(MAIN_NS, "sheetData"))
    children = list(ws_root)
    insert_index = len(children)
    for idx, child in enumerate(children):
        local = child.tag.rsplit("}", 1)[-1]
        if local in {"sheetProtection", "protectedRanges", "autoFilter", "sortState", "dataConsolidate"}:
            insert_index = idx
            break
    ws_root.insert(insert_index, sheet_data)
    return sheet_data


def get_or_create_row(sheet_data, row_idx):
    for idx, row in enumerate(sheet_data.findall(qn(MAIN_NS, "row"))):
        current = int(row.attrib.get("r", "0") or 0)
        if current == row_idx:
            return row
        if current > row_idx:
            new_row = ET.Element(qn(MAIN_NS, "row"), {"r": str(row_idx)})
            sheet_data.insert(idx, new_row)
            return new_row
    new_row = ET.Element(qn(MAIN_NS, "row"), {"r": str(row_idx)})
    sheet_data.append(new_row)
    return new_row


def get_or_create_cell(row_el, row_idx, col_idx):
    ref = cell_ref(col_idx, row_idx)
    for idx, cell in enumerate(row_el.findall(qn(MAIN_NS, "c"))):
        current_ref = cell.attrib.get("r")
        if current_ref == ref:
            return cell
        if current_ref and cell_sort_key(current_ref) > (row_idx, col_idx):
            new_cell = ET.Element(qn(MAIN_NS, "c"), {"r": ref})
            row_el.insert(idx, new_cell)
            return new_cell
    new_cell = ET.Element(qn(MAIN_NS, "c"), {"r": ref})
    row_el.append(new_cell)
    return new_cell


def clear_cell_contents(cell):
    for tag in ("f", "v", "is"):
        child = get_direct_child(cell, tag)
        if child is not None:
            cell.remove(child)
    cell.attrib.pop("t", None)


def set_text_preserve_space(text_el, text):
    text_el.text = text
    if text.startswith(" ") or text.endswith(" ") or "\n" in text:
        text_el.attrib[qn(XML_NS, "space")] = "preserve"
    else:
        text_el.attrib.pop(qn(XML_NS, "space"), None)


def excel_serial_datetime(value, date1904=False):
    if isinstance(value, date) and not isinstance(value, datetime):
        value = datetime.combine(value, time())
    base = datetime(1904, 1, 1) if date1904 else datetime(1899, 12, 30)
    delta = value - base
    return delta.days + (delta.seconds + delta.microseconds / 1_000_000) / 86400


def workbook_uses_1904_dates(workbook_root):
    workbook_pr = workbook_root.find(qn(MAIN_NS, "workbookPr"))
    if workbook_pr is None:
        return False
    return workbook_pr.attrib.get("date1904") in {"1", "true", "True"}


def update_worksheet_dimension(ws_root):
    """Keep worksheet dimensions aligned after direct writes."""
    sheet_data = get_direct_child(ws_root, "sheetData")
    refs = []
    if sheet_data is not None:
        for row_el in sheet_data.findall(qn(MAIN_NS, "row")):
            for cell in row_el.findall(qn(MAIN_NS, "c")):
                ref = cell.attrib.get("r")
                if ref:
                    refs.append(ref)

    dimension = get_direct_child(ws_root, "dimension")
    if dimension is None:
        dimension = ET.Element(qn(MAIN_NS, "dimension"))
        ws_root.insert(0, dimension)

    if not refs:
        dimension.attrib["ref"] = "A1"
        return

    bounds = [split_cell_ref(ref) for ref in refs]
    min_col = min(col for col, _row in bounds)
    max_col = max(col for col, _row in bounds)
    min_row = min(row for _col, row in bounds)
    max_row = max(row for _col, row in bounds)
    start = cell_ref(min_col, min_row)
    end = cell_ref(max_col, max_row)
    dimension.attrib["ref"] = start if start == end else f"{start}:{end}"


def set_formula(cell, formula_text):
    clear_cell_contents(cell)
    f = ET.SubElement(cell, qn(MAIN_NS, "f"))
    f.text = formula_text[1:] if formula_text.startswith("=") else formula_text


def set_inline_string(cell, text):
    clear_cell_contents(cell)
    cell.attrib["t"] = "inlineStr"
    is_el = ET.SubElement(cell, qn(MAIN_NS, "is"))
    t_el = ET.SubElement(is_el, qn(MAIN_NS, "t"))
    set_text_preserve_space(t_el, text)


def set_cell_value_direct(cell, value, date1904=False):
    if value is None:
        clear_cell_contents(cell)
        return

    if isinstance(value, str) and value.startswith("="):
        set_formula(cell, value)
        return

    clear_cell_contents(cell)
    if isinstance(value, bool):
        cell.attrib["t"] = "b"
        ET.SubElement(cell, qn(MAIN_NS, "v")).text = "1" if value else "0"
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.attrib.pop("t", None)
        ET.SubElement(cell, qn(MAIN_NS, "v")).text = repr(value)
    elif isinstance(value, (datetime, date)):
        cell.attrib.pop("t", None)
        ET.SubElement(cell, qn(MAIN_NS, "v")).text = repr(excel_serial_datetime(value, date1904=date1904))
    else:
        set_inline_string(cell, str(value))


def set_workbook_recalc_flags(workbook_root):
    calc_pr = workbook_root.find(qn(MAIN_NS, "calcPr"))
    if calc_pr is None:
        calc_pr = ET.SubElement(workbook_root, qn(MAIN_NS, "calcPr"))
    calc_pr.attrib["calcMode"] = "auto"
    calc_pr.attrib["fullCalcOnLoad"] = "1"
    calc_pr.attrib["forceFullCalc"] = "1"


def value_contains_formula(value):
    if isinstance(value, str):
        return value.startswith("=")
    if isinstance(value, list):
        return any(value_contains_formula(item) for item in value)
    return False


def load_edits_from_args(args):
    if args.edits and args.edits_file:
        print("Error: Provide --edits or --edits-file, not both")
        sys.exit(1)
    if getattr(args, "clear", False) and args.value is not None:
        print("Error: Provide --clear or --value, not both")
        sys.exit(1)
    if args.edits_file:
        with open(args.edits_file, encoding="utf-8-sig") as handle:
            return json.load(handle)
    if args.edits:
        return json.loads(args.edits)
    if (args.cell or getattr(args, "range", None)) and (
        args.value is not None or getattr(args, "clear", False)
    ):
        edit = {
            "value": None if getattr(args, "clear", False) else args.value,
            "type": args.value_type,
        }
        if args.cell:
            edit["cell"] = args.cell
        else:
            edit["range"] = args.range
        return [edit]
    print("Error: Provide --cell/--value, --range/--value, --clear, --edits JSON, or --edits-file")
    sys.exit(1)


def normalise_direct_edit(edit, default_sheet):
    sheet_name = edit.get("sheet", default_sheet)
    if not sheet_name:
        raise ValueError("--sheet is required unless each edit includes a sheet")
    cell = edit.get("cell") or edit.get("range")
    if not cell:
        raise ValueError(f"Edit is missing cell/range: {edit}")
    value = edit.get("values", edit.get("value"))
    value = coerce_edit_value(value, edit.get("type"))
    return sheet_name, cell, value


def direct_edit_workbook(filepath, edits, default_sheet=None, allow_protected=False):
    if find_open_book(filepath)[1] is not None:
        raise RuntimeError(
            "Workbook is open in Excel. Use --engine com to edit the live workbook, "
            "or close Excel before using the direct OOXML engine."
    )

    with zipfile.ZipFile(filepath, "r") as zf:
        workbook_root, rels_root, _sheets_el, sheets = get_workbook_package_map(zf)
        date1904 = workbook_uses_1904_dates(workbook_root)
        grouped = {}
        for edit in edits:
            sheet_name, target, value = normalise_direct_edit(edit, default_sheet)
            if sheet_name not in sheets:
                raise RuntimeError(f"Sheet '{sheet_name}' not found")
            grouped.setdefault(sheet_name, []).append((target, value))

        updates = {}
        formula_written = False
        for sheet_name, sheet_edits in grouped.items():
            part = sheets[sheet_name]["part"]
            ws_root = parse_xml_part(zf, part)
            if not allow_protected and ws_root.find(qn(MAIN_NS, "sheetProtection")) is not None:
                raise RuntimeError(
                    f"Sheet '{sheet_name}' is protected. Use --engine com for Excel's "
                    "locked-cell preflight, or --allow-protected with the direct engine."
                )
            sheet_data = get_or_create_sheet_data(ws_root)
            for target, value in sheet_edits:
                for row_idx, col_idx, item in iter_target_cells(target, value):
                    row_el = get_or_create_row(sheet_data, row_idx)
                    cell = get_or_create_cell(row_el, row_idx, col_idx)
                    if isinstance(item, str) and item.startswith("="):
                        formula_written = True
                    set_cell_value_direct(cell, item, date1904=date1904)
            update_worksheet_dimension(ws_root)
            updates[part] = xml_bytes(ws_root)

        if formula_written:
            set_workbook_recalc_flags(workbook_root)
            removed_calc_chain_parts = remove_calc_chain_relationships(rels_root)
            if removed_calc_chain_parts:
                content_types_root = parse_xml_part(zf, "[Content_Types].xml")
                for part in removed_calc_chain_parts:
                    remove_content_type_override(content_types_root, part)
                updates["xl/_rels/workbook.xml.rels"] = xml_bytes(rels_root)
                updates["[Content_Types].xml"] = xml_bytes(content_types_root)
            updates["xl/workbook.xml"] = xml_bytes(workbook_root)

    rewrite_zip_parts(filepath, updates, deletes=removed_calc_chain_parts if formula_written else None)
    return grouped


def next_relationship_id(rels_root):
    used = {rel.attrib.get("Id", "") for rel in rels_root.findall(qn(PKG_REL_NS, "Relationship"))}
    max_number = 0
    for rid in used:
        match = re.match(r"rId(\d+)$", rid)
        if match:
            max_number = max(max_number, int(match.group(1)))
    candidate = max_number + 1
    while f"rId{candidate}" in used:
        candidate += 1
    return f"rId{candidate}"


def next_sheet_part(zf):
    existing = set(zf.namelist())
    index = 1
    while f"xl/worksheets/sheet{index}.xml" in existing:
        index += 1
    return f"xl/worksheets/sheet{index}.xml"


def next_sheet_id(sheets_el):
    max_id = 0
    for sheet in sheets_el.findall(qn(MAIN_NS, "sheet")):
        try:
            max_id = max(max_id, int(sheet.attrib.get("sheetId", "0")))
        except ValueError:
            pass
    return str(max_id + 1)


def sheet_names_in_order(sheets_el):
    return [
        sheet.attrib.get("name", "")
        for sheet in sheets_el.findall(qn(MAIN_NS, "sheet"))
    ]


def update_sheet_index_references(workbook_root, old_order, new_order):
    """Keep workbook-local indexes aligned after direct sheet reorder/insert."""
    if old_order == new_order:
        return

    new_indexes = {name: index for index, name in enumerate(new_order)}
    old_to_new = {
        index: new_indexes[name]
        for index, name in enumerate(old_order)
        if name in new_indexes
    }

    defined_names = workbook_root.find(qn(MAIN_NS, "definedNames"))
    if defined_names is not None:
        for defined_name in defined_names.findall(qn(MAIN_NS, "definedName")):
            local_sheet_id = defined_name.attrib.get("localSheetId")
            if local_sheet_id is None:
                continue
            try:
                old_index = int(local_sheet_id)
            except ValueError:
                continue
            if old_index in old_to_new:
                defined_name.attrib["localSheetId"] = str(old_to_new[old_index])

    book_views = workbook_root.find(qn(MAIN_NS, "bookViews"))
    if book_views is not None:
        for view in book_views.findall(qn(MAIN_NS, "workbookView")):
            for attr in ("activeTab", "firstSheet"):
                value = view.attrib.get(attr)
                if value is None:
                    continue
                try:
                    old_index = int(value)
                except ValueError:
                    continue
                if old_index in old_to_new:
                    view.attrib[attr] = str(old_to_new[old_index])


def validate_new_sheet_name(name):
    if not name:
        raise RuntimeError("Sheet name cannot be blank")
    if len(name) > 31:
        raise RuntimeError("Sheet name cannot be longer than 31 characters")
    invalid = set("[]:*?/\\")
    bad_chars = sorted({char for char in name if char in invalid})
    if bad_chars:
        raise RuntimeError(f"Sheet name contains invalid character(s): {''.join(bad_chars)}")
    if name.startswith("'") or name.endswith("'"):
        raise RuntimeError("Sheet name cannot start or end with an apostrophe")


def minimal_worksheet_xml():
    ws = ET.Element(qn(MAIN_NS, "worksheet"))
    ET.SubElement(ws, qn(MAIN_NS, "dimension"), {"ref": "A1"})
    ET.SubElement(ET.SubElement(ws, qn(MAIN_NS, "sheetViews")), qn(MAIN_NS, "sheetView"), {"workbookViewId": "0"})
    ET.SubElement(ws, qn(MAIN_NS, "sheetFormatPr"), {"defaultRowHeight": "15"})
    ET.SubElement(ws, qn(MAIN_NS, "sheetData"))
    return xml_bytes(ws)


def ensure_content_type_override(content_types_root, part_name, content_type):
    part_attr = "/" + part_name.lstrip("/")
    for override in content_types_root.findall(qn(CONTENT_TYPES_NS, "Override")):
        if override.attrib.get("PartName") == part_attr:
            override.attrib["ContentType"] = content_type
            return
    ET.SubElement(
        content_types_root,
        qn(CONTENT_TYPES_NS, "Override"),
        {"PartName": part_attr, "ContentType": content_type},
    )


def remove_content_type_override(content_types_root, part_name):
    part_attr = "/" + part_name.lstrip("/")
    for override in list(content_types_root.findall(qn(CONTENT_TYPES_NS, "Override"))):
        if override.attrib.get("PartName") == part_attr:
            content_types_root.remove(override)


def remove_calc_chain_relationships(rels_root):
    removed_parts = []
    for rel in list(rels_root.findall(qn(PKG_REL_NS, "Relationship"))):
        if rel.attrib.get("Type") != CALC_CHAIN_REL_TYPE:
            continue
        target = rel.attrib.get("Target")
        if target:
            removed_parts.append(resolve_relationship_target("xl/workbook.xml", target))
        rels_root.remove(rel)
    return removed_parts


def insert_sheet_element(sheets_el, sheet_el, before=None, after=None):
    sheets = sheets_el.findall(qn(MAIN_NS, "sheet"))
    if before:
        for idx, current in enumerate(sheets):
            if current.attrib.get("name") == before:
                sheets_el.insert(idx, sheet_el)
                return
        raise RuntimeError(f"Insertion sheet '{before}' not found")
    if after:
        for idx, current in enumerate(sheets):
            if current.attrib.get("name") == after:
                sheets_el.insert(idx + 1, sheet_el)
                return
        raise RuntimeError(f"Insertion sheet '{after}' not found")
    sheets_el.append(sheet_el)


def direct_add_sheet(filepath, name, before=None, after=None):
    if find_open_book(filepath)[1] is not None:
        raise RuntimeError("Workbook is open in Excel. Close it before direct sheet changes.")
    validate_new_sheet_name(name)
    with zipfile.ZipFile(filepath, "r") as zf:
        workbook_root, rels_root, sheets_el, sheets = get_workbook_package_map(zf)
        if any(existing.lower() == name.lower() for existing in sheets):
            raise RuntimeError(f"Sheet '{name}' already exists")
        old_order = sheet_names_in_order(sheets_el)
        content_types_root = parse_xml_part(zf, "[Content_Types].xml")
        part = next_sheet_part(zf)
        rid = next_relationship_id(rels_root)
        rel_target = relationship_target_for_part("xl/workbook.xml", part)
        ET.SubElement(
            rels_root,
            qn(PKG_REL_NS, "Relationship"),
            {
                "Id": rid,
                "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
                "Target": rel_target,
            },
        )
        sheet_el = ET.Element(
            qn(MAIN_NS, "sheet"),
            {"name": name, "sheetId": next_sheet_id(sheets_el), qn(REL_NS, "id"): rid},
        )
        insert_sheet_element(sheets_el, sheet_el, before=before, after=after)
        update_sheet_index_references(workbook_root, old_order, sheet_names_in_order(sheets_el))
        ensure_content_type_override(
            content_types_root,
            part,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
        )
        updates = {
            "xl/workbook.xml": xml_bytes(workbook_root),
            "xl/_rels/workbook.xml.rels": xml_bytes(rels_root),
            "[Content_Types].xml": xml_bytes(content_types_root),
            part: minimal_worksheet_xml(),
        }
    rewrite_zip_parts(filepath, updates)
    return part


def direct_move_sheet(filepath, name, before=None, after=None):
    if find_open_book(filepath)[1] is not None:
        raise RuntimeError("Workbook is open in Excel. Close it before direct sheet changes.")
    if bool(before) == bool(after):
        raise RuntimeError("Provide exactly one of --before or --after")
    if before == name or after == name:
        return
    with zipfile.ZipFile(filepath, "r") as zf:
        workbook_root, _rels_root, sheets_el, _sheets = get_workbook_package_map(zf)
        old_order = sheet_names_in_order(sheets_el)
        target_el = None
        for sheet in sheets_el.findall(qn(MAIN_NS, "sheet")):
            if sheet.attrib.get("name") == name:
                target_el = sheet
                break
        if target_el is None:
            raise RuntimeError(f"Sheet '{name}' not found")
        sheets_el.remove(target_el)
        insert_sheet_element(sheets_el, target_el, before=before, after=after)
        update_sheet_index_references(workbook_root, old_order, sheet_names_in_order(sheets_el))
        updates = {"xl/workbook.xml": xml_bytes(workbook_root)}
    rewrite_zip_parts(filepath, updates)


def cmd_read(args):
    filepath = os.path.abspath(args.file)
    if args.backend == "openpyxl":
        wb = load_readonly_workbook(filepath)
        try:
            sheet_name = args.sheet or wb.sheetnames[0]
            if sheet_name not in wb.sheetnames:
                print(f"Error: Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
                sys.exit(1)

            ws = wb[sheet_name]
            print(f"Sheet: {sheet_name}")
            print(f"Dimensions: {openpyxl_dimensions(ws)}")
            print()

            rows = openpyxl_values_as_rows(ws, args.range)
            rows = compact_display_rows(
                rows,
                show_empty_rows=args.show_empty_rows,
                show_empty_cols=args.show_empty_cols,
            )
            if not rows:
                print("(empty sheet)")
                return

            print_markdown_table(rows)
        finally:
            wb.close()
        return

    app, wb, close_on_exit = open_book(
        filepath,
        readonly=True,
        prefer_open=not args.no_attach,
    )

    try:
        sheet_name = args.sheet or wb.sheets[0].name
        names = [s.name for s in wb.sheets]
        if sheet_name not in names:
            print(f"Error: Sheet '{sheet_name}' not found. Available: {', '.join(names)}")
            sys.exit(1)

        ws = wb.sheets[sheet_name]
        used = ws.used_range
        print(f"Sheet: {sheet_name}")
        print(f"Dimensions: {used.address}")
        print()

        if args.range:
            data_range = ws.range(args.range)
        else:
            data_range = used

        raw = range_values_as_rows(data_range)
        raw = compact_display_rows(
            raw,
            show_empty_rows=args.show_empty_rows,
            show_empty_cols=args.show_empty_cols,
        )
        if not raw:
            print("(empty sheet)")
            return

        # Normalise: single cell to nested list, single row to list of lists.
        print_markdown_table(raw)

    finally:
        close_book(app, wb, close_on_exit=close_on_exit)


def cmd_list_sheets(args):
    filepath = os.path.abspath(args.file)
    if args.backend == "openpyxl":
        wb = load_readonly_workbook(filepath)
        try:
            for i, sheet_name in enumerate(wb.sheetnames):
                print(f"  {i + 1}. {sheet_name}")
        finally:
            wb.close()
        return

    app, wb, close_on_exit = open_book(
        filepath,
        readonly=True,
        prefer_open=not args.no_attach,
    )
    try:
        for i, sheet in enumerate(wb.sheets):
            print(f"  {i + 1}. {sheet.name}")
    finally:
        close_book(app, wb, close_on_exit=close_on_exit)


def cmd_search(args):
    filepath = os.path.abspath(args.file)
    if args.backend == "openpyxl":
        wb = load_readonly_workbook(filepath)
        query = args.query.lower()
        try:
            if args.sheet and args.sheet not in wb.sheetnames:
                print(f"Error: Sheet '{args.sheet}' not found. Available: {', '.join(wb.sheetnames)}")
                sys.exit(1)

            sheets = [wb[args.sheet]] if args.sheet else [wb[name] for name in wb.sheetnames]
            found = 0

            for ws in sheets:
                rows = openpyxl_values_as_rows(ws)
                for r_idx, row in enumerate(rows, start=1):
                    for c_idx, val in enumerate(row, start=1):
                        if val is not None and query in str(val).lower():
                            cell = f"{get_column_letter(c_idx)}{r_idx}"
                            print(f"  [{ws.title}] {cell}: {format_cell_value(val)}")
                            found += 1

            if found == 0:
                print(f"  No matches for '{args.query}'")
            else:
                print(f"\n  {found} match(es) found.")
        finally:
            wb.close()
        return

    app, wb, close_on_exit = open_book(
        filepath,
        readonly=True,
        prefer_open=not args.no_attach,
    )
    query = args.query.lower()

    try:
        sheets = [wb.sheets[args.sheet]] if args.sheet else wb.sheets
        found = 0

        for ws in sheets:
            used = ws.used_range
            raw = range_values_as_rows(used)
            if not raw:
                continue

            start_row = used.row
            start_col = used.column

            for r_idx, row in enumerate(raw):
                for c_idx, val in enumerate(row):
                    if val is not None and query in str(val).lower():
                        # Convert to cell address
                        cell = ws.range(start_row + r_idx, start_col + c_idx)
                        print(f"  [{ws.name}] {cell.address.replace('$', '')}: {format_cell_value(val)}")
                        found += 1

        if found == 0:
            print(f"  No matches for '{args.query}'")
        else:
            print(f"\n  {found} match(es) found.")

    finally:
        close_book(app, wb, close_on_exit=close_on_exit)


def cmd_edit(args):
    filepath = os.path.abspath(args.file)
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    edits = load_edits_from_args(args)

    sheet_name = args.sheet
    if not sheet_name and any("sheet" not in edit for edit in edits):
        print("Error: --sheet is required unless each edit includes a sheet")
        sys.exit(1)

    if args.engine == "direct":
        try:
            grouped = direct_edit_workbook(
                filepath,
                edits,
                default_sheet=sheet_name,
                allow_protected=args.allow_protected,
            )
            total = sum(len(sheet_edits) for sheet_edits in grouped.values())
            for edited_sheet, sheet_edits in grouped.items():
                print(f"  Edited {len(sheet_edits)} target(s) on {edited_sheet}")
            print(f"\nSaved via direct OOXML engine: {filepath}")
            if any(value_contains_formula(edit.get("values", edit.get("value"))) for edit in edits):
                print("Formula edits marked workbook for recalculation on next Excel open.")
            print(f"Total target range(s): {total}")
            return
        except Exception as e:
            print(f"Error: {e}")
            sys.exit(1)

    app, wb, close_on_exit = open_book(
        filepath,
        prefer_open=not args.no_attach,
        keep_open=args.keep_open,
    )
    try:
        sheets = {}

        for edit in edits:
            edit_sheet_name = edit.get("sheet", sheet_name)
            if edit_sheet_name not in sheets:
                sheets[edit_sheet_name] = wb.sheets[edit_sheet_name]
            ws = sheets[edit_sheet_name]
            target_ref = edit.get("cell") or edit.get("range")
            if not target_ref:
                raise RuntimeError(f"Edit is missing cell/range: {edit}")
            value = edit.get("values", edit.get("value"))
            value = coerce_edit_value(value, edit.get("type"))
            target = ws.range(target_ref)
            assert_editable_range(ws, target, target_ref)

            if value is None:
                clear_contents_com(target)
                print(f"  Cleared {edit_sheet_name}!{target_ref}")
            else:
                target.value = value
                print(f"  Set {edit_sheet_name}!{target_ref} = {value}")

        close_book(app, wb, save=True, close_on_exit=close_on_exit)
        print(f"\nSaved: {filepath}")

    except Exception as e:
        print(f"Error: {e}")
        close_book(app, wb, close_on_exit=close_on_exit)
        sys.exit(1)


def parse_sheet_mapping(spec):
    """Return (source_sheet, target_sheet) for a copy-sheets argument."""
    if "=" in spec:
        source_name, target_name = spec.split("=", 1)
        return source_name, target_name
    return spec, spec


def cmd_copy_sheets(args):
    filepath = os.path.abspath(args.file)
    source_path = os.path.abspath(args.source)
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    if not os.path.exists(source_path):
        print(f"Error: Source file not found: {source_path}")
        sys.exit(1)

    app, wb, close_on_exit = open_book(
        filepath,
        prefer_open=not args.no_attach,
        keep_open=args.keep_open,
    )
    source_wb = None
    try:
        try:
            app.display_alerts = False
            app.api.DisplayAlerts = False
        except Exception:
            pass

        source_wb = app.books.open(source_path, update_links=False, read_only=True)
        target_names = [sheet.name for sheet in wb.sheets]
        source_names = [sheet.name for sheet in source_wb.sheets]

        if args.after:
            if args.after not in target_names:
                print(f"Error: Target insertion sheet '{args.after}' not found")
                sys.exit(1)
            after_api = wb.sheets[args.after].api
        else:
            after_api = wb.sheets[-1].api

        for spec in args.sheets:
            source_name, target_name = parse_sheet_mapping(spec)
            if source_name not in source_names:
                print(f"Error: Source sheet '{source_name}' not found")
                sys.exit(1)

            if target_name in [sheet.name for sheet in wb.sheets]:
                if not args.replace:
                    print(f"Error: Target sheet '{target_name}' already exists; use --replace")
                    sys.exit(1)
                wb.sheets[target_name].delete()

            after_index = after_api.Index
            sheet_count = wb.api.Worksheets.Count
            source_wb.sheets[source_name].api.Copy(After=after_api)
            if wb.api.Worksheets.Count != sheet_count + 1:
                raise RuntimeError(f"Copy did not add a sheet for '{source_name}'")
            copied_api = wb.api.Worksheets(after_index + 1)
            if copied_api.Name != target_name:
                copied_api.Name = target_name
            after_api = copied_api
            print(f"  Copied {source_name} -> {target_name}")

        if source_wb is not None:
            source_wb.close()
            source_wb = None
        close_book(app, wb, save=True, close_on_exit=close_on_exit)
        print(f"\nSaved: {filepath}")

    except Exception as e:
        print(f"Error: {e}")
        if source_wb is not None:
            try:
                source_wb.close()
            except Exception:
                pass
        close_book(app, wb, close_on_exit=close_on_exit)
        sys.exit(1)


def cmd_add_sheet(args):
    filepath = os.path.abspath(args.file)
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    if args.before and args.after:
        print("Error: Provide --before or --after, not both")
        sys.exit(1)

    try:
        if args.engine == "direct":
            part = direct_add_sheet(filepath, args.name, before=args.before, after=args.after)
            print(f"  Added sheet {args.name} ({part})")
            print(f"\nSaved via direct OOXML engine: {filepath}")
            return

        app, wb, close_on_exit = open_book(
            filepath,
            prefer_open=not args.no_attach,
            keep_open=args.keep_open,
        )
        try:
            names = [sheet.name for sheet in wb.sheets]
            if args.name in names:
                raise RuntimeError(f"Sheet '{args.name}' already exists")
            if args.before:
                new_sheet = wb.sheets.add(args.name, before=wb.sheets[args.before])
            elif args.after:
                new_sheet = wb.sheets.add(args.name, after=wb.sheets[args.after])
            else:
                new_sheet = wb.sheets.add(args.name, after=wb.sheets[-1])
            close_book(app, wb, save=True, close_on_exit=close_on_exit)
            print(f"  Added sheet {new_sheet.name}")
            print(f"\nSaved: {filepath}")
        except Exception:
            close_book(app, wb, close_on_exit=close_on_exit)
            raise
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


def cmd_move_sheet(args):
    filepath = os.path.abspath(args.file)
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)
    if bool(args.before) == bool(args.after):
        print("Error: Provide exactly one of --before or --after")
        sys.exit(1)

    try:
        if args.engine == "direct":
            direct_move_sheet(filepath, args.name, before=args.before, after=args.after)
            print(f"  Moved sheet {args.name}")
            print(f"\nSaved via direct OOXML engine: {filepath}")
            return

        app, wb, close_on_exit = open_book(
            filepath,
            prefer_open=not args.no_attach,
            keep_open=args.keep_open,
        )
        try:
            if args.before:
                wb.sheets[args.name].api.Move(Before=wb.sheets[args.before].api)
            else:
                wb.sheets[args.name].api.Move(After=wb.sheets[args.after].api)
            close_book(app, wb, save=True, close_on_exit=close_on_exit)
            print(f"  Moved sheet {args.name}")
            print(f"\nSaved: {filepath}")
        except Exception:
            close_book(app, wb, close_on_exit=close_on_exit)
            raise
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


def cmd_insert(args):
    filepath = os.path.abspath(args.file)
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    app, wb, close_on_exit = open_book(
        filepath,
        prefer_open=not args.no_attach,
        keep_open=args.keep_open,
    )
    try:
        ws = wb.sheets[args.sheet]
        target = ws.range(args.range)
        if args.axis == "columns":
            target.api.EntireColumn.Insert()
        else:
            target.api.EntireRow.Insert()
        close_book(app, wb, save=True, close_on_exit=close_on_exit)
        print(f"  Inserted {args.axis} at {args.sheet}!{args.range}")
        print(f"\nSaved: {filepath}")
    except Exception as e:
        print(f"Error: {e}")
        close_book(app, wb, close_on_exit=close_on_exit)
        sys.exit(1)


def excel_error_text(value):
    if value is None:
        return None
    text = str(value)
    if text.startswith("#") and any(
        text.startswith(prefix)
        for prefix in ("#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#REF!", "#VALUE!", "#SPILL!", "#CALC!")
    ):
        return text
    return None


def cmd_validate(args):
    filepath = os.path.abspath(args.file)
    if not os.path.exists(filepath):
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    app, wb, close_on_exit = open_book(
        filepath,
        readonly=not args.save,
        prefer_open=not args.no_attach,
        keep_open=args.keep_open,
    )
    try:
        if args.full_calc:
            try:
                app.api.CalculateFullRebuild()
            except Exception:
                app.calculate()
        else:
            app.calculate()

        print(f"Opened and calculated: {filepath}")
        print(f"Sheets: {len(wb.sheets)}")

        if args.check_errors:
            found = []
            for ws in wb.sheets:
                used = ws.used_range
                rows = range_values_as_rows(used)
                start_row = used.row
                start_col = used.column
                for r_idx, row in enumerate(rows):
                    for c_idx, value in enumerate(row):
                        error = excel_error_text(value)
                        if error:
                            addr = ws.range(start_row + r_idx, start_col + c_idx).address.replace("$", "")
                            found.append((ws.name, addr, error))
                            if len(found) >= args.max_errors:
                                break
                    if len(found) >= args.max_errors:
                        break
                if len(found) >= args.max_errors:
                    break
            if found:
                print("Formula/value errors:")
                for sheet, addr, error in found:
                    print(f"  [{sheet}] {addr}: {error}")
                if len(found) >= args.max_errors:
                    print(f"  Stopped after {args.max_errors} errors.")
            else:
                print("No visible Excel error values found in used ranges.")

        close_book(app, wb, save=args.save, close_on_exit=close_on_exit)
        if args.save:
            print("Saved recalculated workbook.")
    except Exception as e:
        print(f"Error: {e}")
        close_book(app, wb, close_on_exit=close_on_exit)
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="Read and edit Excel workbooks")
    sub = parser.add_subparsers(dest="command")

    def add_attach_options(command_parser):
        command_parser.add_argument(
            "--no-attach",
            action="store_true",
            help="With --backend com, do not attach to an already-open workbook; open a temporary copy instead",
        )

    def add_read_backend_option(command_parser):
        command_parser.add_argument(
            "--backend",
            choices=("openpyxl", "com"),
            default="openpyxl",
            help="Read backend. Default is openpyxl because it never saves the workbook.",
        )

    # read
    p_read = sub.add_parser("read", help="Read a sheet as a markdown table")
    p_read.add_argument("file")
    p_read.add_argument("--sheet", help="Sheet name (default: first sheet)")
    p_read.add_argument("--range", help="Cell range e.g. A1:D10")
    p_read.add_argument(
        "--show-empty-rows",
        action="store_true",
        help="Preserve fully empty rows in display output",
    )
    p_read.add_argument(
        "--show-empty-cols",
        action="store_true",
        help="Preserve fully empty columns in display output",
    )
    add_read_backend_option(p_read)
    add_attach_options(p_read)

    # list-sheets
    p_list = sub.add_parser("list-sheets", help="List sheet names")
    p_list.add_argument("file")
    add_read_backend_option(p_list)
    add_attach_options(p_list)

    # search
    p_search = sub.add_parser("search", help="Search for text across sheets")
    p_search.add_argument("file")
    p_search.add_argument("--query", "-q", required=True)
    p_search.add_argument("--sheet", help="Limit to a specific sheet")
    add_read_backend_option(p_search)
    add_attach_options(p_search)

    # edit
    p_edit = sub.add_parser("edit", help="Edit cells (preserves all formatting)")
    p_edit.add_argument("file")
    p_edit.add_argument("--sheet")
    p_edit.add_argument("--cell", help="Cell reference e.g. A1")
    p_edit.add_argument("--range", help="Cell range e.g. A1:D10 for scalar or matrix writes")
    p_edit.add_argument("--value", help="New value for the cell")
    p_edit.add_argument("--clear", action="store_true", help="Clear the target cell or range")
    p_edit.add_argument("--edits", help='Batch JSON: [{"cell":"A1","value":"x"}, ...]')
    p_edit.add_argument("--edits-file", help="Path to a JSON file containing batch edits")
    p_edit.add_argument(
        "--engine",
        choices=("direct", "com"),
        default="direct",
        help="Write engine. direct edits OOXML parts without Excel; com edits through Excel.",
    )
    p_edit.add_argument(
        "--allow-protected",
        action="store_true",
        help="Allow direct OOXML writes to protected sheets. Default is to fail on protected sheets.",
    )
    p_edit.add_argument(
        "--value-type",
        choices=("text", "string", "date", "datetime", "int", "integer", "float", "number", "numeric"),
        help="Optional type coercion for --cell/--value edits",
    )
    p_edit.add_argument(
        "--keep-open",
        action="store_true",
        help="With --engine com, if the workbook is not already open, open it visibly and leave it open",
    )
    add_attach_options(p_edit)

    # copy-sheets
    p_copy = sub.add_parser("copy-sheets", help="Copy sheets from another workbook through Excel COM")
    p_copy.add_argument("file", help="Target workbook")
    p_copy.add_argument("--source", required=True, help="Source workbook")
    p_copy.add_argument(
        "--sheets",
        nargs="+",
        required=True,
        help='Sheet names to copy. Use "Source=Target" to rename a copied sheet.',
    )
    p_copy.add_argument("--after", help="Target sheet to insert after (default: end of workbook)")
    p_copy.add_argument("--replace", action="store_true", help="Replace target sheets with the same name")
    p_copy.add_argument(
        "--keep-open",
        action="store_true",
        help="If the target workbook is not already open, open it visibly and leave it open",
    )
    add_attach_options(p_copy)

    # add-sheet
    p_add_sheet = sub.add_parser("add-sheet", help="Add a blank worksheet")
    p_add_sheet.add_argument("file")
    p_add_sheet.add_argument("name")
    p_add_sheet.add_argument("--before", help="Insert before this sheet")
    p_add_sheet.add_argument("--after", help="Insert after this sheet")
    p_add_sheet.add_argument(
        "--engine",
        choices=("direct", "com"),
        default="direct",
        help="Sheet creation engine. direct edits workbook XML; com uses Excel.",
    )
    p_add_sheet.add_argument(
        "--keep-open",
        action="store_true",
        help="With --engine com, leave the workbook open",
    )
    add_attach_options(p_add_sheet)

    # move-sheet
    p_move_sheet = sub.add_parser("move-sheet", help="Move a worksheet before or after another sheet")
    p_move_sheet.add_argument("file")
    p_move_sheet.add_argument("name")
    p_move_sheet.add_argument("--before", help="Move before this sheet")
    p_move_sheet.add_argument("--after", help="Move after this sheet")
    p_move_sheet.add_argument(
        "--engine",
        choices=("direct", "com"),
        default="direct",
        help="Move engine. direct edits workbook XML; com uses Excel.",
    )
    p_move_sheet.add_argument(
        "--keep-open",
        action="store_true",
        help="With --engine com, leave the workbook open",
    )
    add_attach_options(p_move_sheet)

    # insert
    p_insert = sub.add_parser("insert", help="Insert rows or columns through Excel COM")
    p_insert.add_argument("file")
    p_insert.add_argument("--sheet", required=True)
    p_insert.add_argument("--axis", choices=("rows", "columns"), required=True)
    p_insert.add_argument("--range", required=True, help="Rows/columns/range to insert, e.g. D:D or 12:12")
    p_insert.add_argument(
        "--keep-open",
        action="store_true",
        help="Leave the workbook open after inserting",
    )
    add_attach_options(p_insert)

    # validate
    p_validate = sub.add_parser("validate", help="Open in Excel, calculate, and optionally save/check errors")
    p_validate.add_argument("file")
    p_validate.add_argument("--full-calc", action="store_true", help="Run Excel CalculateFullRebuild")
    p_validate.add_argument("--save", action="store_true", help="Save after calculation")
    p_validate.add_argument("--check-errors", action="store_true", help="Scan used ranges for visible Excel error values")
    p_validate.add_argument("--max-errors", type=int, default=50, help="Maximum formula/value errors to report")
    p_validate.add_argument(
        "--keep-open",
        action="store_true",
        help="Leave the workbook open after validation",
    )
    add_attach_options(p_validate)

    args = parser.parse_args()

    commands = {
        "read": cmd_read,
        "list-sheets": cmd_list_sheets,
        "search": cmd_search,
        "edit": cmd_edit,
        "copy-sheets": cmd_copy_sheets,
        "add-sheet": cmd_add_sheet,
        "move-sheet": cmd_move_sheet,
        "insert": cmd_insert,
        "validate": cmd_validate,
    }

    if args.command in commands:
        commands[args.command](args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
